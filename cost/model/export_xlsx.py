"""Выгрузка экономики блока в xlsx — для заказчика, которому привычен Excel.

Файл не воспроизводит смету: это те же четыре представления, что и на
вкладке, — итоги, структура, натуральные показатели и параметры прогона.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from cost.model.inputs import BlockEconomics
from cost.v2.models import CostLayer


LAYER_LABELS: dict[str, str] = {
    CostLayer.VARIABLE.value: "Переменные затраты",
    CostLayer.PROJECT_DIRECT.value: "Прямые затраты блока",
    CostLayer.PRODUCTION.value: "Производственная себестоимость",
    CostLayer.FULL.value: "Полная себестоимость",
}

PRICE_LABELS: tuple[tuple[str, str], ...] = (
    ("marginal", "Маржинальная цена, ₽/м³"),
    ("full", "Полная себестоимость, ₽/м³"),
    ("with_margin", "Цена с ОХР и рентабельностью, ₽/м³"),
    ("with_vat", "Цена с НДС, ₽/м³"),
)


def build_workbook(
    economics: BlockEconomics,
    *,
    passport_name: str,
    parameters: Mapping[str, Any],
) -> Workbook:
    workbook = Workbook()
    _totals_sheet(workbook.active, economics, passport_name)
    _structure_sheet(workbook.create_sheet("Структура"), economics)
    _natural_sheet(workbook.create_sheet("Натуральные показатели"), economics)
    _parameters_sheet(workbook.create_sheet("Параметры"), parameters, economics)
    return workbook


def export_bytes(
    economics: BlockEconomics,
    *,
    passport_name: str,
    parameters: Mapping[str, Any],
) -> bytes:
    stream = BytesIO()
    build_workbook(economics, passport_name=passport_name, parameters=parameters).save(stream)
    return stream.getvalue()


def _totals_sheet(sheet: Any, economics: BlockEconomics, passport_name: str) -> None:
    sheet.title = "Итоги"
    _title(sheet, f"Экономика блока: {passport_name}")
    rows: list[tuple[str, Any]] = [("Объём блока, м³", float(economics.block_volume_m3))]
    for layer, value in economics.layer_totals.items():
        rows.append((LAYER_LABELS.get(layer.value, layer.value) + ", ₽", float(value)))
    for key, label in PRICE_LABELS:
        rows.append((label, float(economics.price_per_m3.get(key, Decimal("0")))))
    markup = economics.markup
    for key, label in (
        ("overhead_rub", "Общехозяйственные расходы, ₽"),
        ("margin_rub", "Рентабельность, ₽"),
        ("price_rub", "Цена блока без НДС, ₽"),
        ("vat_rub", "НДС, ₽"),
        ("price_with_vat_rub", "Цена блока с НДС, ₽"),
    ):
        if key in markup:
            rows.append((label, float(markup[key])))
    _table(sheet, ("Показатель", "Значение"), rows, start_row=3)


def _structure_sheet(sheet: Any, economics: BlockEconomics) -> None:
    rows = [
        (
            LAYER_LABELS.get(line.layer.value, line.layer.value),
            line.operation_code,
            line.cost_item_code,
            line.cost_item_name,
            float(line.amount_rub),
            line.formula,
        )
        for line in economics.lines
    ]
    _title(sheet, "Структура затрат по слоям")
    _table(
        sheet,
        ("Слой", "Операция", "Код статьи", "Статья", "Сумма, ₽", "Формула"),
        rows,
        start_row=3,
    )


def _natural_sheet(sheet: Any, economics: BlockEconomics) -> None:
    natural = economics.natural
    rows = [
        (key, float(value), natural.lineage.get(key, ""))
        for key, value in sorted(natural.values.items())
    ]
    _title(sheet, "Натуральные показатели блока")
    _table(sheet, ("Величина", "Значение", "Происхождение"), rows, start_row=3)


def _parameters_sheet(
    sheet: Any, parameters: Mapping[str, Any], economics: BlockEconomics
) -> None:
    rows = [(str(key), _text(value)) for key, value in parameters.items()]
    rows.extend(("Предупреждение", warning) for warning in economics.warnings)
    _title(sheet, "Параметры прогона")
    _table(sheet, ("Параметр", "Значение"), rows, start_row=3)


def _title(sheet: Any, text: str) -> None:
    sheet["A1"] = text
    sheet["A1"].font = Font(bold=True, size=13)


def _table(
    sheet: Any, header: Sequence[str], rows: Sequence[Sequence[Any]], *, start_row: int
) -> None:
    for column, name in enumerate(header, start=1):
        cell = sheet.cell(row=start_row, column=column, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for offset, row in enumerate(rows, start=start_row + 1):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=offset, column=column, value=value)
    for column, name in enumerate(header, start=1):
        width = max(len(str(name)), *(len(str(row[column - 1])) for row in rows)) if rows else len(str(name))
        sheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 12), 60)


def _text(value: Any) -> str:
    if isinstance(value, (list, tuple)):
        return "; ".join(_text(item) for item in value)
    if isinstance(value, dict):
        return "; ".join(f"{key}={_text(item)}" for key, item in value.items())
    return "" if value is None else str(value)
