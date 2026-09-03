"""Справочники файлом: книга xlsx (лист на раздел) и JSON-снимок.

Состав колонок берётся из JSON-схемы раздела, поэтому новый раздел или поле
попадают в файл без правок здесь. Модуль не знает ни об API, ни об
интерфейсе: принимает и возвращает `ReferenceSnapshot` / `ReferenceItem`.
"""
from __future__ import annotations

import io
import json
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable, Mapping

from openpyxl import Workbook, load_workbook

from cost.v2.models import ReferenceItem, ReferenceSnapshot
from cost.v2.references import REFERENCE_SECTION_DEFINITIONS
from cost.v2.schemas import SECTION_SCHEMAS, section_json_schema

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
JSON_MEDIA_TYPE = "application/json"

TRUE_WORDS = frozenset({"да", "true", "1", "yes", "истина"})
FALSE_WORDS = frozenset({"нет", "false", "0", "no", "ложь"})


class ReferenceFileError(ValueError):
    """Файл не разобран: сообщение называет лист, строку и колонку."""


@dataclass(frozen=True)
class Column:
    key: str
    title: str
    kind: str  # text | bool | date | decimal | json


FIXED_COLUMNS: tuple[Column, ...] = (
    Column("code", "Код", "text"),
    Column("name", "Наименование", "text"),
    Column("is_active", "Активна", "bool"),
    Column("valid_from", "Действует с", "date"),
    Column("valid_to", "Действует по", "date"),
    Column("comment", "Комментарий", "text"),
)
_FIXED_KEYS = {column.key for column in FIXED_COLUMNS}


def exportable_sections() -> list[str]:
    """Разделы со схемой в порядке каталога; устаревшие в файл не идут."""

    return [
        code
        for code, meta in REFERENCE_SECTION_DEFINITIONS.items()
        if code in SECTION_SCHEMAS and not meta.get("deprecated", False)
    ]


def section_columns(section: str) -> list[Column]:
    schema = section_json_schema(section)
    columns = list(FIXED_COLUMNS)
    for key, prop in schema.get("properties", {}).items():
        if prop.get("x-internal") or key in _FIXED_KEYS:
            continue
        columns.append(Column(key, _title(key, prop), _kind(prop)))
    return columns


def _title(key: str, prop: Mapping[str, Any]) -> str:
    title = str(prop.get("title") or prop.get("description") or key)
    unit = prop.get("x-unit")
    return f"{title}, {unit}" if unit else title


def _types(prop: Mapping[str, Any]) -> set[str]:
    found: set[str] = set()
    if "type" in prop:
        found.add(str(prop["type"]))
    for variant in prop.get("anyOf", ()):
        if "type" in variant:
            found.add(str(variant["type"]))
        if "$ref" in variant or "items" in variant:
            found.add("object")
    if "$ref" in prop or "items" in prop:
        found.add("object")
    return found


def _kind(prop: Mapping[str, Any]) -> str:
    if "x-ref" in prop or "enum" in prop:
        return "text"
    types = _types(prop)
    if "boolean" in types:
        return "bool"
    if "array" in types or "object" in types:
        return "json"
    if "number" in types or "integer" in types:
        return "decimal"
    if prop.get("format") == "date":
        return "date"
    return "text"


# --- Экспорт ----------------------------------------------------------------


def export_xlsx(snapshot: ReferenceSnapshot) -> bytes:
    book = Workbook()
    book.remove(book.active)
    for section in exportable_sections():
        sheet = book.create_sheet(section)
        columns = section_columns(section)
        sheet.append([column.key for column in columns])
        sheet.append([column.title for column in columns])
        for item in snapshot.sections.get(section, ()):
            sheet.append([_cell(column, _item_value(item, column.key)) for column in columns])
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def export_json(snapshot: ReferenceSnapshot) -> dict[str, Any]:
    data = snapshot.to_dict()
    known = set(exportable_sections())
    data["sections"] = {key: value for key, value in data["sections"].items() if key in known}
    return data


def _item_value(item: ReferenceItem, key: str) -> Any:
    if key in _FIXED_KEYS:
        return getattr(item, key)
    return item.payload.get(key)


def _cell(column: Column, value: Any) -> Any:
    if value is None or value == "":
        return None
    if column.kind == "bool":
        return "да" if bool(value) else "нет"
    if column.kind == "date":
        return value if isinstance(value, date) else date.fromisoformat(str(value))
    if column.kind == "decimal":
        try:
            number = Decimal(str(value))
        except InvalidOperation:
            return str(value)
        return int(number) if number == number.to_integral_value() else float(number)
    if column.kind == "json":
        if value in ([], {}):
            return None
        return json.dumps(value, ensure_ascii=False)
    return str(value)


# --- Импорт -----------------------------------------------------------------


def import_file(name: str, data: bytes) -> dict[str, list[ReferenceItem]]:
    lowered = name.lower()
    if lowered.endswith(".xlsx"):
        return import_xlsx(data)
    if lowered.endswith(".json"):
        return import_json(data)
    raise ReferenceFileError("Поддерживаются файлы xlsx или JSON.")


def import_json(data: bytes) -> dict[str, list[ReferenceItem]]:
    try:
        payload = json.loads(data.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ReferenceFileError(f"Файл не является корректным JSON: {exc}") from exc
    raw_sections = payload.get("sections") if isinstance(payload, dict) else None
    if not isinstance(raw_sections, dict):
        raise ReferenceFileError("В JSON нет объекта «sections» с разделами.")
    known = set(exportable_sections())
    sections: dict[str, list[ReferenceItem]] = {}
    for section, rows in raw_sections.items():
        if section not in known:
            raise ReferenceFileError(f"Неизвестный раздел «{section}».")
        if not isinstance(rows, list):
            raise ReferenceFileError(f"Раздел «{section}» должен быть списком записей.")
        sections[section] = [ReferenceItem.from_dict(row) for row in rows]
    return sections


def import_xlsx(data: bytes) -> dict[str, list[ReferenceItem]]:
    try:
        book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl бросает разные классы для битых файлов
        raise ReferenceFileError(f"Файл не открывается как книга xlsx: {exc}") from exc
    known = set(exportable_sections())
    sections: dict[str, list[ReferenceItem]] = {}
    for sheet in book.worksheets:
        section = sheet.title
        if section not in known:
            raise ReferenceFileError(f"Неизвестный лист «{section}»: такого раздела нет.")
        columns = {column.key: column for column in section_columns(section)}
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            sections[section] = []
            continue
        keys: list[str | None] = [str(key).strip() if key not in (None, "") else None for key in header]
        for key in keys:
            if key is not None and key not in columns:
                raise ReferenceFileError(f"Неизвестная колонка «{key}» (лист «{section}»).")
        next(rows, None)  # строка подписей
        items: list[ReferenceItem] = []
        for row_no, values in enumerate(rows, start=3):
            cells = {key: value for key, value in zip(keys, values) if key is not None}
            if all(value in (None, "") for value in cells.values()):
                continue
            items.append(_item_from_cells(section, row_no, cells, columns))
        sections[section] = items
    return sections


def _item_from_cells(
    section: str,
    row_no: int,
    cells: Mapping[str, Any],
    columns: Mapping[str, Column],
) -> ReferenceItem:
    def parse(key: str) -> Any:
        return _parse(section, row_no, columns[key], cells.get(key))

    code = parse("code") if "code" in cells else None
    if not code:
        raise ReferenceFileError(f"Лист «{section}», строка {row_no}: нет кода записи.")
    name = parse("name") if "name" in cells else None
    if not name:
        raise ReferenceFileError(f"Лист «{section}», строка {row_no}: нет наименования записи.")
    is_active = parse("is_active") if "is_active" in cells else None
    payload: dict[str, Any] = {}
    for key in cells:
        if key in _FIXED_KEYS:
            continue
        value = parse(key)
        if value is not None:
            payload[key] = value
    return ReferenceItem(
        code=str(code),
        name=str(name),
        payload=payload,
        is_active=True if is_active is None else bool(is_active),
        valid_from=parse("valid_from") if "valid_from" in cells else None,
        valid_to=parse("valid_to") if "valid_to" in cells else None,
        source="Импорт xlsx",
        comment=str(parse("comment") or "") if "comment" in cells else "",
    )


def _parse(section: str, row_no: int, column: Column, value: Any) -> Any:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    where = f"лист «{section}», строка {row_no}, колонка «{column.key}»"
    if column.kind == "bool":
        if isinstance(value, bool):
            return value
        word = str(value).strip().lower()
        if word in TRUE_WORDS:
            return True
        if word in FALSE_WORDS:
            return False
        raise ReferenceFileError(f"{where}: ожидается «да» или «нет», получено «{value}».")
    if column.kind == "date":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            return date.fromisoformat(str(value).strip())
        except ValueError as exc:
            raise ReferenceFileError(f"{where}: ожидается дата, получено «{value}».") from exc
    if column.kind == "decimal":
        try:
            number = Decimal(str(value).strip().replace(",", "."))
        except InvalidOperation as exc:
            raise ReferenceFileError(f"{where}: ожидается число, получено «{value}».") from exc
        return format(number.normalize(), "f")
    if column.kind == "json":
        if isinstance(value, (list, dict)):
            return value
        try:
            return json.loads(str(value))
        except json.JSONDecodeError as exc:
            raise ReferenceFileError(f"{where}: ожидается JSON, получено «{value}».") from exc
    return str(value).strip()
