"""Render released mass-blast document contexts into PDF, XLSX and ZIP bytes."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile


DOCUMENT_TITLES = {
    "PROJECT": "Проект массового взрыва",
    "ORDER": "Распорядок массового взрыва",
    "SCHEDULE": "График заряжания и производства взрыва",
}


def _title(kind: str) -> str:
    return DOCUMENT_TITLES.get(kind, "Документ массового взрыва")


def _pdf_text(value: object) -> str:
    """User and reference data must not become ReportLab paragraph markup."""

    return escape(str(value if value not in (None, "") else "—"))


def _font_path() -> str:
    candidates = (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    )
    for candidate in candidates:
        if Path(candidate).exists():
            return candidate
    raise RuntimeError("Не найден шрифт с кириллицей для PDF. Установите fonts-dejavu-core.")


def render_pdf(context: dict, kind: str = "PROJECT") -> bytes:
    """Create a print-ready Russian PDF without reading any source workbook."""

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    except ImportError as exc:  # pragma: no cover - exercised by container build
        raise RuntimeError("Для выпуска PDF требуется пакет reportlab.") from exc

    font_name = "BlastEXDejaVu"
    if font_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(font_name, _font_path()))
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    heading = ParagraphStyle("blast-heading", parent=styles["Heading1"], fontName=font_name, fontSize=15, leading=18, spaceAfter=10)
    body = ParagraphStyle("blast-body", parent=styles["BodyText"], fontName=font_name, fontSize=8.5, leading=11)
    small = ParagraphStyle("blast-small", parent=body, fontSize=7, leading=8)
    project = context.get("project", {})
    totals = context.get("totals", {})
    story = [Paragraph(_title(kind), heading)]
    story.append(Paragraph(f"Объект работ: {_pdf_text(project.get('object_name'))}", body))
    story.append(Paragraph(f"Код объекта: {_pdf_text(project.get('site_code'))} · дата: {_pdf_text(project.get('blast_date'))} {_pdf_text(project.get('blast_time'))}", body))
    story.append(Paragraph(f"Заказчик: {_pdf_text(project.get('customer_code'))} · профиль: {_pdf_text(project.get('document_profile_code') or 'STANDARD')}", body))
    story.append(Spacer(1, 5 * mm))
    summary = [
        ["Показатель", "Значение"],
        ["Количество блоков", str(totals.get("block_count", 0))],
        ["Количество скважин", str(totals.get("hole_count", 0))],
        ["Объём горной массы, м³", f"{float(totals.get('block_volume_m3', 0)):,.3f}"],
        ["Метры бурения, м", f"{float(totals.get('drilling_m', 0)):,.3f}"],
        ["Масса ВМ, кг", f"{float(totals.get('explosive_mass_kg', 0)):,.3f}"],
        ["Удельный расход, кг/м³", f"{float(totals.get('specific_q_kg_m3', 0)):,.6f}"],
        ["Макс. масса на замедление, кг", f"{float(totals.get('max_charge_per_delay_kg', 0)):,.3f}"],
    ]
    table = Table(summary, colWidths=[105 * mm, 65 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name), ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E7F1EC")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B5C5BC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([table, Spacer(1, 6 * mm)])
    story.append(Paragraph("Параметры скважин", heading))
    rows = [["Блок", "Скважина", "Ø, мм", "Глубина, м", "Заряд, кг", "Забойка, м", "Боёвики, шт."]]
    for block in context.get("blocks", []):
        for hole in (block.get("snapshot") or {}).get("holes", []):
            rows.append([
                str(block.get("code") or "—"), str(hole.get("hole_id") or "—"), str(hole.get("diameter_mm") or 0),
                f"{float(hole.get('depth_m', 0)):.2f}", f"{float(hole.get('charge_kg', 0)):.2f}",
                f"{float(hole.get('stemming_m', 0)):.2f}", str(hole.get("primer_count") or 0),
            ])
    holes_table = Table(rows, colWidths=[23 * mm, 28 * mm, 18 * mm, 25 * mm, 25 * mm, 25 * mm, 26 * mm], repeatRows=1)
    holes_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name), ("FONTSIZE", (0, 0), (-1, -1), 6.4),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E7F1EC")),
        ("GRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#B5C5BC")),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.extend([holes_table, Spacer(1, 5 * mm)])
    story.append(Paragraph("Ответственные лица", heading))
    responsibility_rows = [["Роль", "Сотрудник", "Должность"]] + [
        [str(item.get("role_code", "")), str(item.get("employee_name", "")), str(item.get("position_name", ""))]
        for item in context.get("responsibilities", [])
    ]
    responsibility_table = Table(responsibility_rows, colWidths=[55 * mm, 60 * mm, 60 * mm], repeatRows=1)
    responsibility_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name), ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E7F1EC")), ("GRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#B5C5BC")),
    ]))
    story.extend([responsibility_table, Spacer(1, 4 * mm)])
    safety = context.get("safety_plan") or {}
    story.append(Paragraph(f"Опасная зона: {_pdf_text(safety.get('danger_zone_radius_m'))} м. Сигнальный профиль: {_pdf_text((context.get('signal_plan') or {}).get('profile_code'))}.", body))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("Документ сформирован из неизменяемой ревизии BlastEX. Технические параметры, цена и себестоимость не смешиваются.", small))
    doc.build(story)
    return output.getvalue()


def render_xlsx(context: dict, kind: str = "PROJECT") -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:  # pragma: no cover - exercised by container build
        raise RuntimeError("Для выпуска XLSX требуется пакет openpyxl.") from exc

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Проект МВ"
    project = context.get("project", {})
    totals = context.get("totals", {})
    summary.append([_title(kind)])
    summary.append(["Объект работ", project.get("object_name", "")])
    summary.append(["Код объекта", project.get("site_code", "")])
    summary.append(["Дата взрыва", project.get("blast_date", "")])
    summary.append(["Количество блоков", totals.get("block_count", 0)])
    summary.append(["Количество скважин", totals.get("hole_count", 0)])
    summary.append(["Объём горной массы, м³", totals.get("block_volume_m3", 0)])
    summary.append(["Метры бурения, м", totals.get("drilling_m", 0)])
    summary.append(["Масса ВМ, кг", totals.get("explosive_mass_kg", 0)])
    summary.append(["Удельный расход, кг/м³", totals.get("specific_q_kg_m3", 0)])
    summary.merge_cells("A1:B1")
    summary["A1"].font = Font(bold=True, size=14)
    for row in summary.iter_rows(min_row=2, max_col=2):
        row[0].font = Font(bold=True)
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 26

    holes = workbook.create_sheet("Скважины")
    headers = ["Блок", "Скважина", "Ряд", "Колонка", "Тип", "Ø, мм", "Глубина, м", "Перебур, м", "Обводнение", "Заряд, кг", "ВВ", "Забойка, м", "Боёвики, шт."]
    holes.append(headers)
    for block in context.get("blocks", []):
        for hole in (block.get("snapshot") or {}).get("holes", []):
            holes.append([
                block.get("code", ""), hole.get("hole_id", ""), hole.get("row", 0), hole.get("col", 0), hole.get("kind", ""),
                hole.get("diameter_mm", 0), hole.get("depth_m", 0), hole.get("subdrill_m", 0), "да" if hole.get("watered") else "нет",
                hole.get("charge_kg", 0), hole.get("charge_product", ""), hole.get("stemming_m", 0), hole.get("primer_count", 0),
            ])
    fill = PatternFill("solid", fgColor="E7F1EC")
    for cell in holes[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True)
    holes.freeze_panes = "A2"
    for column in "ABCDEFGHIJKLM":
        holes.column_dimensions[column].width = 16
    holes.column_dimensions["A"].width = 24
    holes.column_dimensions["B"].width = 18
    holes.column_dimensions["K"].width = 28

    responsibility = workbook.create_sheet("Ответственные")
    responsibility.append(["Роль", "Сотрудник", "Код сотрудника", "Должность"])
    for item in context.get("responsibilities", []):
        responsibility.append([item.get("role_code", ""), item.get("employee_name", ""), item.get("employee_code", ""), item.get("position_name", "")])
    for cell in responsibility[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for column in "ABCD":
        responsibility.column_dimensions[column].width = 30

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def render_zip(context: dict, project_pdf: bytes, project_xlsx: bytes, revision_label: str) -> bytes:
    output = BytesIO()
    with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr(f"{revision_label}_проект_МВ.pdf", project_pdf)
        archive.writestr(f"{revision_label}_проект_МВ.xlsx", project_xlsx)
        archive.writestr("manifest.json", __import__("json").dumps({"context_version": context.get("document_context_version"), "formula_version": context.get("formula_version")}, ensure_ascii=False, indent=2))
    return output.getvalue()
