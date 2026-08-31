"""Mass-blast project context is a snapshot, never a second BVR engine."""
from __future__ import annotations

import copy
from io import BytesIO
import unittest

from design.mass_blast import (
    block_from_design,
    build_document_context,
    content_sha256,
    has_blocking_issues,
    validate_project_context,
)
from tests.scenario_fixtures import charged_design


class MassBlastContextTests(unittest.TestCase):
    def _context(self):
        design = charged_design("mass-blast")
        design.design_id = "design-mb-01"
        design.revision = 4
        design.designed_sha256 = "a" * 64
        block = block_from_design(design, code="Блок 14", horizon="+120")
        payload = {
            "name": "Массовый взрыв №14",
            "site_code": "SITE-14",
            "object_name": "Карьер Заготовка",
            "customer_code": "CUSTOMER-1",
            "blast_date": "2026-08-31",
            "blast_time": "14:00",
            "document_profile_code": "STANDARD",
            "reference_revision_id": "ref-1",
            "responsibilities": [
                {"role_code": "blast_manager", "employee_code": "E-1", "employee_name": "Руководитель"},
                {"role_code": "explosives_supervisor", "employee_code": "E-2", "employee_name": "Ответственный"},
            ],
            "safety_plan": {"danger_zone_radius_m": 500},
            "signal_plan": {"profile_code": "THREE_SIGNALS"},
            "guard_posts": [{"code": "P-1", "location": "Северный пост"}],
            "notifications": [{"recipient": "Диспетчер"}],
        }
        return design, block, build_document_context(payload, [block])

    def test_context_uses_design_snapshot_and_has_aggregates(self):
        design, block, context = self._context()
        self.assertEqual(context["totals"]["block_count"], 1)
        self.assertEqual(context["totals"]["hole_count"], len([hole for hole in design.holes if hole.enabled]))
        self.assertGreater(context["totals"]["explosive_mass_kg"], 0)
        self.assertEqual(context["blocks"][0]["design_sha256"], "a" * 64)
        self.assertEqual(block.snapshot["lineage"]["technical"], "DESIGN")
        self.assertNotIn("cost", context)
        self.assertNotIn("market_price", context)

    def test_validation_requires_safety_and_responsibilities(self):
        _, _, context = self._context()
        self.assertFalse(has_blocking_issues(validate_project_context(context)))
        incomplete = copy.deepcopy(context)
        incomplete["responsibilities"] = []
        incomplete["safety_plan"] = {}
        incomplete["guard_posts"] = []
        incomplete["signal_plan"] = {}
        issues = validate_project_context(incomplete)
        self.assertTrue(has_blocking_issues(issues))
        self.assertEqual({issue.code for issue in issues if issue.level == "error"}, {
            "responsibility_required", "danger_zone", "signal_profile", "guard_posts"
        })

    def test_release_requires_a_current_graphic_attachment_when_profile_demands_it(self):
        _, _, context = self._context()
        issues = validate_project_context(context, require_attachments=True)
        self.assertIn("attachments", {issue.code for issue in issues if issue.level == "error"})
        context["attachments"] = [{"id": "file-1", "sha256": "b" * 64}]
        self.assertFalse(has_blocking_issues(validate_project_context(context, require_attachments=True)))

    def test_hash_is_reproducible_and_changes_when_project_changes(self):
        _, _, context = self._context()
        first = content_sha256(context)
        self.assertEqual(first, content_sha256(copy.deepcopy(context)))
        changed = copy.deepcopy(context)
        changed["project"]["blast_date"] = "2026-09-01"
        self.assertNotEqual(first, content_sha256(changed))

    def test_rendered_xlsx_and_pdf_are_real_standalone_artifacts(self):
        try:
            from openpyxl import load_workbook
            from design.mass_blast_rendering import render_pdf, render_xlsx
        except ImportError:
            self.skipTest("Document renderer dependencies are installed in the API image.")
        _, _, context = self._context()
        workbook = load_workbook(BytesIO(render_xlsx(context)))
        self.assertEqual(workbook.sheetnames, ["Проект МВ", "Скважины", "Ответственные"])
        self.assertGreater(workbook["Скважины"].max_row, 1)
        self.assertTrue(render_pdf(context).startswith(b"%PDF-"))


if __name__ == "__main__":
    unittest.main()
