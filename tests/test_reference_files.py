"""Экспорт и импорт справочников файлом (спецификация §8)."""
from __future__ import annotations

import io
import json
from datetime import date

import pytest
from openpyxl import load_workbook

from cost.v2.models import ReferenceItem, ReferenceSnapshot
from cost.v2.references import default_reference_snapshot
from cost.v2.schemas import SECTION_SCHEMAS


def _snapshot(**sections) -> ReferenceSnapshot:
    base = dict(default_reference_snapshot().sections)
    base.update({key: tuple(items) for key, items in sections.items()})
    return ReferenceSnapshot(revision_id="REV-FILE", sections=base, published_by="tester")


def _site() -> ReferenceItem:
    return ReferenceItem(
        code="SITE_A",
        name="Карьер А",
        payload={"mobilization_km": "220", "is_watered": True, "customer_code": "CP_1"},
        valid_from=date(2026, 1, 1),
        comment="тест",
    )


def _crew() -> ReferenceItem:
    return ReferenceItem(
        code="CREW_1",
        name="Бригада",
        payload={"package_code": "PKG_DRILL", "members": [{"position_code": "POS_DRILLER", "headcount": "2"}]},
    )


class TestColumns:
    def test_fixed_columns_come_first_and_internal_fields_are_hidden(self):
        from cost.v2.reference_files import FIXED_COLUMNS, section_columns

        columns = section_columns("sites")
        assert [c.key for c in columns[: len(FIXED_COLUMNS)]] == [
            "code", "name", "is_active", "valid_from", "valid_to", "comment",
        ]
        keys = [c.key for c in columns]
        assert "legacy_ref" not in keys
        assert "mobilization_km" in keys and "is_watered" in keys and "customer_code" in keys

    def test_kinds_follow_the_schema(self):
        from cost.v2.reference_files import section_columns

        kinds = {c.key: c.kind for c in section_columns("sites")}
        assert kinds["mobilization_km"] == "decimal"
        assert kinds["is_watered"] == "bool"
        assert kinds["customer_code"] == "text"
        assert kinds["valid_from"] == "date"
        crew = {c.key: c.kind for c in section_columns("crew_templates")}
        assert crew["members"] == "json"
        materials = {c.key: c.kind for c in section_columns("materials")}
        assert materials["storage_class"] == "text"

    def test_titles_carry_units(self):
        from cost.v2.reference_files import section_columns

        titles = {c.key: c.title for c in section_columns("sites")}
        assert titles["mobilization_km"] == "Плечо мобилизации, км"
        assert titles["code"] == "Код"

    def test_every_schema_section_is_exportable(self):
        from cost.v2.reference_files import exportable_sections

        assert set(exportable_sections()) == set(SECTION_SCHEMAS)
        assert all(len(name) <= 31 for name in exportable_sections())


class TestExport:
    def test_xlsx_has_a_sheet_per_section_with_keys_titles_and_rows(self):
        from cost.v2.reference_files import export_xlsx

        book = load_workbook(io.BytesIO(export_xlsx(_snapshot(sites=[_site()], crew_templates=[_crew()]))))
        assert "sites" in book.sheetnames and "crew_templates" in book.sheetnames
        sheet = book["sites"]
        keys = [cell.value for cell in sheet[1]]
        titles = [cell.value for cell in sheet[2]]
        assert keys[:6] == ["code", "name", "is_active", "valid_from", "valid_to", "comment"]
        assert titles[:2] == ["Код", "Наименование"]
        row = {key: cell.value for key, cell in zip(keys, sheet[3])}
        assert row["code"] == "SITE_A"
        assert row["name"] == "Карьер А"
        assert row["is_active"] == "да"
        assert row["is_watered"] == "да"
        assert row["valid_from"] == date(2026, 1, 1) or getattr(row["valid_from"], "date", lambda: None)() == date(2026, 1, 1)
        assert row["mobilization_km"] == 220
        assert row["customer_code"] == "CP_1"
        assert row["comment"] == "тест"
        crew = book["crew_templates"]
        crew_keys = [cell.value for cell in crew[1]]
        crew_row = {key: cell.value for key, cell in zip(crew_keys, crew[3])}
        assert json.loads(crew_row["members"]) == [{"position_code": "POS_DRILLER", "headcount": "2"}]

    def test_empty_section_still_has_header_rows(self):
        from cost.v2.reference_files import export_xlsx

        book = load_workbook(io.BytesIO(export_xlsx(_snapshot())))
        sheet = book["sites"]
        assert sheet.max_row == 2
        assert sheet["A1"].value == "code"

    def test_json_export_is_the_snapshot_object(self):
        from cost.v2.reference_files import export_json

        payload = export_json(_snapshot(sites=[_site()]))
        assert payload["revision_id"] == "REV-FILE"
        assert payload["published_by"] == "tester"
        assert payload["sections"]["sites"][0]["code"] == "SITE_A"
        assert payload["sections"]["sites"][0]["valid_from"] == "2026-01-01"
        assert "section_catalog" not in payload


class TestImport:
    def test_xlsx_round_trip_restores_items(self):
        from cost.v2.reference_files import export_xlsx, import_xlsx

        sections = import_xlsx(export_xlsx(_snapshot(sites=[_site()], crew_templates=[_crew()])))
        site = sections["sites"][0]
        assert site.code == "SITE_A" and site.name == "Карьер А"
        assert site.is_active is True
        assert site.valid_from == date(2026, 1, 1) and site.valid_to is None
        assert site.comment == "тест"
        assert site.source == "Импорт xlsx"
        assert site.payload["mobilization_km"] == "220"
        assert site.payload["is_watered"] is True
        assert site.payload["customer_code"] == "CP_1"
        assert "legacy_ref" not in site.payload
        assert "distance_from_base_km" not in site.payload
        crew = sections["crew_templates"][0]
        assert crew.payload["members"] == [{"position_code": "POS_DRILLER", "headcount": "2"}]
        assert sections["rocks"] == []

    def test_round_trip_of_the_default_snapshot_validates(self):
        from cost.v2.reference_files import export_xlsx, import_xlsx
        from cost.v2.references import has_validation_errors, validate_reference_sections

        snapshot = default_reference_snapshot()
        sections = import_xlsx(export_xlsx(snapshot))
        assert not has_validation_errors(validate_reference_sections(sections))
        assert {i.code for i in sections["units"]} == {i.code for i in snapshot.sections["units"]}
        assert [i.code for i in sections["operations"]] == [i.code for i in snapshot.sections["operations"]]

    def test_unknown_sheet_and_column_are_reported_by_name(self):
        from cost.v2.reference_files import ReferenceFileError, export_xlsx, import_xlsx

        book = load_workbook(io.BytesIO(export_xlsx(_snapshot())))
        book.create_sheet("прочее").append(["code", "name"])
        with pytest.raises(ReferenceFileError, match="лист «прочее»"):
            import_xlsx(_bytes(book))

        book = load_workbook(io.BytesIO(export_xlsx(_snapshot())))
        book["sites"].cell(row=1, column=book["sites"].max_column + 1, value="mobilisation")
        with pytest.raises(ReferenceFileError, match="колонка «mobilisation».*лист «sites»"):
            import_xlsx(_bytes(book))

    def test_bad_boolean_names_sheet_row_and_column(self):
        from cost.v2.reference_files import ReferenceFileError, export_xlsx, import_xlsx

        book = load_workbook(io.BytesIO(export_xlsx(_snapshot(sites=[_site()]))))
        sheet = book["sites"]
        keys = [cell.value for cell in sheet[1]]
        sheet.cell(row=3, column=keys.index("is_watered") + 1, value="возможно")
        with pytest.raises(ReferenceFileError, match="лист «sites», строка 3, колонка «is_watered»"):
            import_xlsx(_bytes(book))

    def test_row_without_code_is_an_error_but_blank_rows_are_skipped(self):
        from cost.v2.reference_files import ReferenceFileError, export_xlsx, import_xlsx

        book = load_workbook(io.BytesIO(export_xlsx(_snapshot(sites=[_site()]))))
        sheet = book["sites"]
        sheet.append([None] * sheet.max_column)
        assert len(import_xlsx(_bytes(book))["sites"]) == 1
        sheet.append([None, "Без кода"])
        with pytest.raises(ReferenceFileError, match="строка 5.*нет кода"):
            import_xlsx(_bytes(book))

    def test_json_import_reads_snapshot_object(self):
        from cost.v2.reference_files import ReferenceFileError, export_json, import_json

        payload = export_json(_snapshot(sites=[_site()]))
        sections = import_json(json.dumps(payload, ensure_ascii=False).encode("utf-8"))
        assert sections["sites"][0].code == "SITE_A"
        assert sections["sites"][0].valid_from == date(2026, 1, 1)
        with pytest.raises(ReferenceFileError, match="раздел «other»"):
            import_json(json.dumps({"sections": {"other": []}}).encode("utf-8"))
        with pytest.raises(ReferenceFileError, match="sections"):
            import_json(b"{}")
        with pytest.raises(ReferenceFileError, match="JSON"):
            import_json(b"not json")

    def test_import_file_dispatches_by_extension(self):
        from cost.v2.reference_files import ReferenceFileError, export_json, export_xlsx, import_file

        assert "sites" in import_file("refs.XLSX", export_xlsx(_snapshot()))
        assert "sites" in import_file("refs.json", json.dumps(export_json(_snapshot())).encode("utf-8"))
        with pytest.raises(ReferenceFileError, match="xlsx или JSON"):
            import_file("refs.csv", b"")


def _bytes(book) -> bytes:
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()
